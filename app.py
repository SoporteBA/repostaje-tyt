import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from datetime import datetime

# --- 1. CONFIGURACIÓN Y CONSTANTES ---
st.set_page_config(
    page_title="T&T | MOEVE > NOVATRANS",
    page_icon="🚛",
    layout="centered"
)

# Filtro de matrículas
MATRICULAS_EXCLUIDAS = {
    "TJT-001", "TJT-002", "TJT-003", "TJT-004", "TJT-005", "TJT-006", "TJT-007"
}

# Mapeo de conceptos
MAPPING_PRODUCTOS = {
    "DIESEL STAR": "Gasoleo",
    "ECOBLUE": "AdBlue",
    "SIN PLOMO": "Gasoil B",
    "AUTOPISTAS DE PEAJE": "Peaje",
    "GEST. SERV. AUTOP. ESPAÑA": "Otros"
}

# --- 2. ESTILOS CSS ---
st.markdown("""
    <style>
    [data-testid="stFileUploader"] small { display: none; }
    [data-testid="stFileUploaderDropzone"] div div::before { content: "Arrastra y suelta tu archivo aquí"; }
    button[kind="secondary"] { background-color: #fbfbfb; border: 1px solid #d6d6d6; }
    .footer { position: fixed; bottom: 10px; width: 100%; text-align: center; color: #888; font-size: 12px; }
    </style>
""", unsafe_allow_html=True)

# --- 3. LOGO Y CABECERA ---
col_izq, col_cen, col_der = st.columns([1, 2, 1])
with col_cen:
    try:
        st.image("tyt_logo_trans.png", use_container_width=True)
    except Exception:
        pass # Si no hay logo, no mostramos nada ni error

st.markdown("<h1 style='text-align: center;'>Creador de plantilla NOVATRANS</h1>", unsafe_allow_html=True)
st.info("Sube el Excel de Moeve y la Plantilla de Novatrans. El sistema rellenará los datos respetando estrictamente los formatos de la plantilla.")

# --- 4. FUNCIÓN DE PROCESAMIENTO ---
def procesar_archivos(plantilla, datos):
    # A. Leer Datos Origen (Moeve/Cepsa)
    try:
        # Forzamos que 'Tarjeta' sea string para no perder ceros iniciales
        df_origen = pd.read_excel(datos, header=2, dtype={'Tarjeta': str, 'Matricula': str})
    except Exception as e:
        st.error(f"❌ Error leyendo Excel de Moeve: {e}")
        return None

    # B. Leer Plantilla Destino (Novatrans)
    try:
        # load_workbook mantiene los estilos originales del archivo
        wb_destino = load_workbook(plantilla)
        ws_destino = wb_destino.active
    except Exception as e:
        st.error(f"❌ Error leyendo Plantilla Novatrans: {e}")
        return None

    # C. Limpieza de datos previos en la plantilla
    # Solo borramos VALORES, no formatos
    max_row = ws_destino.max_row
    max_col = ws_destino.max_column
    if max_row > 1:
        for row in ws_destino.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.value = None

    # D. Procesamiento fila a fila
    fila_destino = 2
    
    # UI: Barra de progreso
    texto_estado = st.empty()
    barra = st.progress(0)
    total_filas = len(df_origen)

    for index, row in df_origen.iterrows():
        # Actualizar barra
        if index % 10 == 0 or index == total_filas - 1:
            barra.progress((index + 1) / total_filas)
            texto_estado.text(f"Procesando registro {index + 1} de {total_filas}...")

        # 1. Filtro Matrícula
        matricula_actual = str(row.get('Matricula', '')).strip()
        if matricula_actual in MATRICULAS_EXCLUIDAS:
            continue

        # 2. Gestión de Fechas
        # Convertimos a objetos datetime de Python puros.
        # Al pasarlos a Excel, Excel aplicará el formato que tenga la columna configurada.
        fecha_hora_raw = row.get('Fecha y hora')
        val_fecha = None
        val_hora = None
        
        if pd.notnull(fecha_hora_raw):
            try:
                # Intentamos convertir a datetime
                dt_obj = pd.to_datetime(fecha_hora_raw, dayfirst=True)
                val_fecha = dt_obj.date() # Solo la parte fecha
                val_hora = dt_obj.time()  # Solo la parte hora
            except:
                pass # Se queda en None

        # 3. Limpieza Tarjeta
        tarjeta_valor = str(row.get('Tarjeta', ''))
        if tarjeta_valor == 'nan': tarjeta_valor = ""
        if tarjeta_valor.endswith('.0'): tarjeta_valor = tarjeta_valor[:-2]

        # 4. Mapeo Productos
        concepto_original = df_origen.iloc[index, 10]
        nombre_concepto = str(concepto_original).strip() if pd.notnull(concepto_original) else ""
        producto_final = MAPPING_PRODUCTOS.get(nombre_concepto, concepto_original)

        # 5. Escritura en Celdas (SIN TOCAR .number_format)
        
        # Col 2 (B): Matrícula
        ws_destino.cell(row=fila_destino, column=2).value = matricula_actual
        
        # Col 3 (C): Producto
        ws_destino.cell(row=fila_destino, column=3).value = producto_final
        
        # Col 4 (D): Kilómetros / Datos
        ws_destino.cell(row=fila_destino, column=4).value = df_origen.iloc[index, 4]
        
        # Col 5 (E): Tarjeta
        # Insertamos como string. Si la plantilla tiene formato Texto, perfecto.
        # Si tiene formato General, al ser string, Excel suele respetar ceros iniciales si se pasa así.
        ws_destino.cell(row=fila_destino, column=5).value = tarjeta_valor
        
        # Resto de columnas (Importes, Litros, etc.)
        # Asumimos que son números. Pandas los lee como float/int, se pasan igual.
        ws_destino.cell(row=fila_destino, column=6).value = df_origen.iloc[index, 11]
        ws_destino.cell(row=fila_destino, column=7).value = df_origen.iloc[index, 8]
        ws_destino.cell(row=fila_destino, column=8).value = df_origen.iloc[index, 12]
        ws_destino.cell(row=fila_destino, column=9).value = df_origen.iloc[index, 13]

        # Fechas y Horas (Col 12 y 13)
        if val_fecha:
            ws_destino.cell(row=fila_destino, column=12).value = val_fecha # Pasa objeto Date
            
        if val_hora:
            ws_destino.cell(row=fila_destino, column=13).value = val_hora # Pasa objeto Time

        fila_destino += 1

    texto_estado.success("✅ ¡Procesamiento completado!")
    barra.empty()
    
    # Guardar
    output = io.BytesIO()
    wb_destino.save(output)
    output.seek(0)
    return output

# --- 5. INTERFAZ ---
col1, col2 = st.columns(2)
with col1:
    uploaded_plantilla = st.file_uploader("📂 1. Plantilla Novatrans (XLSX)", type=["xlsx"])
with col2:
    uploaded_datos = st.file_uploader("📄 2. Excel Moeve/Cepsa (XLSX)", type=["xlsx"])

st.divider()

if uploaded_plantilla and uploaded_datos:
    if st.button("🚀 Procesar Archivos", type="primary", use_container_width=True):
        with st.spinner('⏳ Insertando datos sin alterar formatos...'):
            archivo_final = procesar_archivos(uploaded_plantilla, uploaded_datos)
            
            if archivo_final:
                st.download_button(
                    label="⬇️ Descargar Resultado",
                    data=archivo_final,
                    file_name=f"Importacion_Novatrans_{datetime.now().strftime('%d-%m-%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
else:
    st.warning("⚠️ Sube ambos archivos para comenzar.")

st.markdown(
    """<div class='footer'>
        Herramienta interna para Tránsitos y Transportes Logísticos | v2.0 Formats Fixed
    </div>""", 
    unsafe_allow_html=True
)
